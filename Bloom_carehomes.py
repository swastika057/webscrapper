import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def load_carehomes_list(driver, url):
    """Load main page and extract list of carehomes with title, location, URL."""
    wait = WebDriverWait(driver, 30)
    driver.get(url)
    container = wait.until(EC.presence_of_element_located(
        (By.ID, "locations-content")))
    blocks = container.find_elements(By.XPATH, ".//div/header/div")

    carehome_data = []
    for block in blocks:
        try:
            a_tag = block.find_element(By.TAG_NAME, "a")
            title = a_tag.get_attribute("title").strip()
            href = a_tag.get_attribute("href").strip()
            p_tag = block.find_element(By.TAG_NAME, "p")
            location = p_tag.text.strip()

            carehome_data.append({
                "Parent_Company": "BloomCare",
                "Business_Name": title,
                "Location": location,
                "URL": href
            })
        except NoSuchElementException:
            continue
    return carehome_data


def extract_manager_name(driver, url):
    """Extract manager name from a carehome detail page."""
    wait = WebDriverWait(driver, 20)
    manager_name = "Manager not found"
    try:
        driver.get(url)
        wait.until(EC.presence_of_element_located(
            (By.CLASS_NAME, "profile-row-section")))

        manager_li_element = wait.until(
            EC.presence_of_element_located(
                (By.XPATH, "//li[div[text()='Person in charge']]"))
        )
        full_name_element = manager_li_element.find_element(
            By.XPATH, "./following-sibling::li[1]")
        full_name = full_name_element.text.strip()
        manager_name = full_name.split('(')[0].strip()
    except (NoSuchElementException, TimeoutException):
        pass
    return manager_name


def extract_contact_number(driver, url):
    """Extract contact number by clicking phone button and 'Something else'."""
    wait = WebDriverWait(driver, 20)
    contact_number = "Contact number not found"
    try:
        driver.get(url)
        wait.until(EC.presence_of_element_located(
            (By.CLASS_NAME, "profile-row-section")))

        # Click "View Phone Number" button
        call_button = wait.until(
            EC.element_to_be_clickable((By.ID, "brochure_phone")))
        driver.execute_script("arguments[0].click();", call_button)
        time.sleep(1)

        # Wait for and click the "Something else" radio label
        wait.until(EC.visibility_of_element_located((By.ID, "phone_modal")))
        label_something_else = wait.until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "label[for='enquiry_type_general']"))
        )
        driver.execute_script("arguments[0].click();", label_something_else)
        time.sleep(2)

        # Extract phone number link text
        wait.until(EC.visibility_of_element_located(
            (By.ID, "telephone_enquiry")))
        phone_link = wait.until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "#telephone_enquiry a[href^='tel:']"))
        )
        contact_number = phone_link.text.strip()

        # Close modal
        try:
            close_button = driver.find_element(
                By.CSS_SELECTOR, "#ajaxModal button.close")
            driver.execute_script("arguments[0].click();", close_button)
            wait.until(EC.invisibility_of_element_located(
                (By.ID, "ajaxModal")))
        except:
            pass
    except:
        pass
    return contact_number


def extract_manager_and_contact(driver, url):
    """Extract both manager name and contact number from carehome page."""
    manager = extract_manager_name(driver, url)
    contact = extract_contact_number(driver, url)
    return manager, contact


def print_carehome_info(data, index, total):
    """Print formatted carehome information."""
    print(
        f"\n--- Processing carehome {index}/{total}: {data['Business_Name']} ---")
    print(f"Parent_Company   : {data['Parent_Company']}")
    print(f"Business_Name    : {data['Business_Name']}")
    print(f"Location         : {data['Location']}")
    print(f"Manager          : {data.get('Manager', 'Manager not found')}")
    print(
        f"Contact_Number   : {data.get('Contact_Number', 'Contact number not found')}")
    print(f"Status           : {data.get('Status','')}")
    print("-" * 60)


def save_to_excel(data, filename):
    """Save list of dict data to Excel file with dropdown in Status column."""
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)
    print(f" Data saved to {filename}")

    wb = load_workbook(filename)
    ws = wb.active

    # Find "Status" column index (1-based)
    header = [cell.value for cell in ws[1]]
    if "Status" in header:
        status_col = header.index("Status") + 1

        dv = DataValidation(
            type="list", formula1='"Success,Pending,Failure"', allow_blank=True)
        ws.add_data_validation(dv)

        for row in range(2, ws.max_row + 1):
            dv.add(ws.cell(row=row, column=status_col))

        wb.save(filename)
        print(f" Dropdown added to 'Status' column in {filename}")
    else:
        print(" 'Status' column not found.")


def main():
    driver = uc.Chrome(headless=False)
    print(" Opening BloomCare group page...")
    try:
        url = "https://www.carehome.co.uk/care_search_results.cfm/searchgroup/65432218733"
        carehomes = load_carehomes_list(driver, url)
        results = []

        total = len(carehomes)
        for i, home in enumerate(carehomes, start=1):
            manager, contact = extract_manager_and_contact(driver, home['URL'])
            home['Manager'] = manager
            home['Contact_Number'] = contact
            home['Status'] = "Pending"  # Set default status
            print_carehome_info(home, i, total)
            results.append(home)
            time.sleep(2)

        save_to_excel(results, "bloomcare_data.xlsx")

    except Exception as e:
        print(f"Error during scraping: {e}")
    finally:
        input("\nPress ENTER to close browser...")
        driver.quit()


if __name__ == "__main__":
    main()
